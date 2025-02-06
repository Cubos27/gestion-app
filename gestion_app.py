import tkinter as tk
import os
from tkinter import ttk, messagebox, Toplevel
from datetime import datetime, timedelta
from openpyxl import Workbook, load_workbook
from tkcalendar import DateEntry
from fpdf import FPDF

class FinanceApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Gestionapp")
        self.root.geometry("1200x700")
        self.root.configure(bg='#f0f0f0')
        
        # Configurar estilo
        self.style = ttk.Style()
        self.style.theme_use('clam')
        self.configure_styles()
        
        # Configurar almacenamiento
        self.filename = "finanzas.xlsx"
        self.transactions = []
        
        # Variables de control
        self.selected_period = tk.StringVar(value="Día")
        self.selected_transaction = None
        
        # Crear interfaz
        self.create_widgets()
        self.load_transactions()
        self.update_table()
        
    def configure_styles(self):
        self.colors = {
            'primary': '#2a73ff',
            'success': '#28a745',
            'danger': '#dc3545',
            'light': '#f8f9fa',
            'dark': '#212529'
        }
        
        # Estilos generales
        self.style.configure('TFrame', background='#f0f0f0')
        self.style.configure('TLabel', background='#f0f0f0', foreground='#212529')
        self.style.configure('Header.TLabel', font=('Arial', 12, 'bold'), foreground='#2a73ff')
        
        # Botones
        self.style.configure('Primary.TButton', 
                           font=('Arial', 10, 'bold'),
                           borderwidth=0,
                           relief='flat',
                           background=self.colors['primary'],
                           foreground='white',
                           padding=10)
        
        self.style.configure('Secondary.TButton', 
                           font=('Arial', 10),
                           borderwidth=0,
                           relief='flat',
                           background=self.colors['light'],
                           foreground=self.colors['dark'],
                           padding=10)
        
        self.style.configure('Success.TButton', 
                           font=('Arial', 10, 'bold'),
                           borderwidth=0,
                           relief='flat',
                           background='#28a745',
                           foreground='white',
                           padding=10)
        
        self.style.map('Primary.TButton',
            background=[('active', '#1a5ee6'), ('disabled', '#cccccc')],
            foreground=[('active', 'white'), ('disabled', '#666666')]
        )
        
        self.style.map('Secondary.TButton',
            background=[('active', '#e9ecef'), ('disabled', '#cccccc')],
            foreground=[('active', self.colors['dark']), ('disabled', '#666666')]
        )

        self.style.map('Success.TButton',
            background=[('active', '#218838'), ('disabled', '#cccccc')],
            foreground=[('active', 'white'), ('disabled', '#666666')]
        )

        # ComboBox
        self.style.configure('TCombobox',
            selectbackground=self.colors['primary'],
            fieldbackground='white',
            background='white',
            bordercolor='#ced4da',
            darkcolor='#ffffff',
            lightcolor='#ffffff',
            arrowsize=12,
            padding=8,
            relief='flat'
        )

        self.style.map('TCombobox',
            fieldbackground=[('readonly', 'white')],
            background=[('readonly', 'white')],
            bordercolor=[('focus', self.colors['primary']), ('!focus', '#ced4da')],
            arrowcolor=[('!disabled', self.colors['dark']), ('disabled', '#868e96')]
        )
        
        # Tabla
        self.style.configure('Treeview.Heading', 
                           font=('Arial', 10, 'bold'), 
                           background='#e9ecef',
                           relief='flat',
                           borderwidth=0)
        
        self.style.configure('Treeview', 
                           rowheight=30,
                           background='white',
                           fieldbackground='white',
                           borderwidth=0)
        
        self.style.map('Treeview',
            background=[('selected', '#e2e5e9')]
        )
        
        # Entradas
        self.style.configure('TEntry',
                           bordercolor='#ced4da',
                           lightcolor='#ffffff',
                           darkcolor='#ffffff',
                           padding=8,
                           relief='flat')
        
        self.style.configure('TCombobox',
                           selectbackground=self.colors['primary'],
                           fieldbackground='white',
                           padding=8,
                           relief='flat')
        
    def create_widgets(self):
        # Frame principal
        main_frame = ttk.Frame(self.root)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=30, pady=30)
        
        # Header
        header_frame = ttk.Frame(main_frame)
        header_frame.pack(fill=tk.X, pady=(0, 20))
        
        ttk.Label(header_frame, text="Gestión Financiera", style='Header.TLabel').pack(side=tk.LEFT)
        
        # Controles
        control_frame = ttk.Frame(header_frame)
        control_frame.pack(side=tk.RIGHT)
        
        # Selector de período
        period_selector = ttk.Combobox(
            control_frame,
            textvariable=self.selected_period,
            values=["Día", "Semana", "Mes", "Mostrar todo"],
            state='readonly',
            width=10,
            style='Custom.TCombobox'
        )
        period_selector.bind('<<ComboboxSelected>>', lambda e: self.update_table())
        period_selector.pack(side=tk.LEFT, padx=10)
        
        # Botones
        btn_frame = ttk.Frame(control_frame)
        btn_frame.pack(side=tk.LEFT)
        
        ttk.Button(btn_frame, text="Nuevo", style='Primary.TButton', command=self.open_add_window).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="Editar", style='Secondary.TButton', command=self.open_edit_window).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="Eliminar", style='Secondary.TButton', command=self.delete_transaction).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="Generar Reporte", style='Success.TButton', command=self.open_balance_window).pack(side=tk.LEFT, padx=5)
        
        # Tabla
        columns = ("Fecha", "Tipo", "Categoría", "Monto", "Descripción")
        self.tree = ttk.Treeview(
            main_frame,
            columns=columns,
            show="headings",
            selectmode="browse"
        )
        
        for col in columns:
            self.tree.heading(col, text=col, anchor=tk.CENTER)
            self.tree.column(col, width=120, anchor=tk.CENTER)
            
        self.tree.column("Descripción", width=300)
        self.tree.pack(fill=tk.BOTH, expand=True)

    def open_balance_window(self):
        self.balance_window = Toplevel(self.root)
        self.balance_window.title("Generar Reporte")
        self.balance_window.geometry("400x300")
        self.balance_window.configure(bg='#f5f6fa')
        
        main_frame = ttk.Frame(self.balance_window)
        main_frame.pack(padx=20, pady=20, fill=tk.BOTH, expand=True)
        
        # Selector de fechas
        ttk.Label(main_frame, text="Fecha Inicio:").pack(pady=5, anchor=tk.W)
        self.start_date = DateEntry(main_frame, 
                                  date_pattern='yyyy-mm-dd',
                                  background='#2a73ff',
                                  foreground='white',
                                  bordercolor='#ced4da')
        self.start_date.pack(fill=tk.X, pady=5)
        
        ttk.Label(main_frame, text="Fecha Fin:").pack(pady=5, anchor=tk.W)
        self.end_date = DateEntry(main_frame, 
                                date_pattern='yyyy-mm-dd',
                                background='#2a73ff',
                                foreground='white',
                                bordercolor='#ced4da')
        self.end_date.pack(fill=tk.X, pady=5)
        
        # Botón generar
        btn_frame = ttk.Frame(main_frame)
        btn_frame.pack(pady=20)
        
        ttk.Button(btn_frame, text="Generar PDF", style='Primary.TButton',
                 command=self.generate_pdf_report).pack(side=tk.LEFT, padx=10)
        
        ttk.Button(btn_frame, text="Cancelar", style='Secondary.TButton',
                 command=self.balance_window.destroy).pack(side=tk.LEFT, padx=10)

    def generate_pdf_report(self):
        try:
            start = datetime.strptime(self.start_date.get(), "%Y-%m-%d")
            end = datetime.strptime(self.end_date.get(), "%Y-%m-%d")
            
            if start > end:
                raise ValueError("La fecha de inicio debe ser anterior a la fecha final")
                
            filtered = [
                t for t in self.transactions
                if start <= datetime.strptime(t["Fecha"], "%Y-%m-%d") <= end
            ]
            
            if not filtered:
                messagebox.showwarning("Advertencia", "No hay transacciones en el rango seleccionado")
                return
                
            # Calcular totales
            ingresos = sum(float(t["Monto"]) for t in filtered if t["Tipo"] == "Ingreso")
            gastos = sum(float(t["Monto"]) for t in filtered if t["Tipo"] == "Gasto")
            ganancia = ingresos - gastos
            margen = (ganancia / ingresos * 100) if ingresos != 0 else 0
            
            # Separar transacciones
            ingresos_lista = [t for t in filtered if t["Tipo"] == "Ingreso"]
            gastos_lista = [t for t in filtered if t["Tipo"] == "Gasto"]

            # Configurar PDF
            pdf = FPDF(orientation='P', unit='mm', format='A4')
            pdf.set_auto_page_break(auto=True, margin=15)
            pdf.add_page()
            
            # Estilos
            pdf.set_draw_color(34, 119, 255)
            pdf.set_line_width(0.3)
            
            # Encabezado 
            pdf.set_font('Arial', 'B', 16)
            pdf.cell(0, 10, 'Reporte Financiero', 0, 1, 'C')
            pdf.set_font('Arial', '', 12)
            pdf.cell(0, 8, f'Periodo: {start.strftime("%d/%m/%Y")} - {end.strftime("%d/%m/%Y")}', 0, 1, 'C')
            pdf.ln(10)

            # Tabla de Ingresos
            if ingresos_lista:
                pdf.set_font('Arial', 'B', 12)
                pdf.cell(0, 8, 'Ingresos', 0, 1)
                self.generar_tabla_con_total(pdf, ingresos_lista, "Ingresos Totales:")
                pdf.ln(8)

            # Tabla de Gastos
            if gastos_lista:
                pdf.set_font('Arial', 'B', 12)
                pdf.cell(0, 8, 'Gastos', 0, 1)
                self.generar_tabla_con_total(pdf, gastos_lista, "Gastos Totales:")
                pdf.ln(10)
            
            # Resumen financiero
            pdf.ln(10)
            pdf.set_font('Arial', 'B', 12)
            pdf.set_fill_color(245, 245, 245)  # Gris claro
            pdf.cell(0, 10, 'Resumen General', 0, 1, 'L')
            
            pdf.set_font('Times', 'B', 11)
            pdf.cell(60, 8, 'Concepto', 1, 0, 'C', True)
            pdf.cell(60, 8, 'Monto', 1, 1, 'C', True)
            
            pdf.set_font('Times', '', 11)
            data = [
                ('Ingresos Totales', f"${ingresos:,.2f}"),
                ('Gastos Totales', f"${gastos:,.2f}"),
                ('Ganancia Neta', f"${ganancia:,.2f}"),
                ('Margen de Ganancia', f"{margen:.1f}%")
            ]
            
            for label, value in data:
                pdf.cell(60, 8, label, 1, 0, 'L')
                pdf.cell(60, 8, value, 1, 1, 'R')
            
            # Pie de página
            pdf.set_y(-20)
            pdf.set_font('Arial', 'I', 8)
            pdf.cell(0, 5, 'Este reporte fue generado automáticamente por gestionapp', 0, 0, 'C')
            # pdf.ln(5)
            # pdf.cell(0, 5, f'Generado el: {datetime.now().strftime("%d/%m/%Y %H:%M")}', 0, 0, 'C')
            
            # Guardar archivo
            if not os.path.exists(os.path.join("Reportes")):
                os.makedirs(os.path.join("Reportes"))

            filename = f"Reporte_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
            pathPDF = os.path.join("Reportes", filename)
            pdf.output(pathPDF)
            
            messagebox.showinfo("Éxito", f"Reporte generado: {filename}")
            self.balance_window.destroy()
            
        except ValueError as e:
            messagebox.showerror("Error", f"Datos inválidos: {str(e)}")
        except Exception as e:
            messagebox.showerror("Error", f"Ocurrió un error: {str(e)}")

    def generar_tabla_con_total(self, pdf, datos, texto_total):
        # Configurar columnas
        columnas = ['Fecha', 'Categoría', 'Descripción', 'Monto']
        anchos = [25, 35, 80, 25]
        
        # Encabezado
        pdf.set_fill_color(240, 245, 255)
        pdf.set_font('Arial', 'B', 10)
        for ancho, col in zip(anchos, columnas):
            pdf.cell(ancho, 8, col, 1, 0, 'C', True)
        pdf.ln()
        
        # Filas
        pdf.set_font('Times', '', 10)
        total = 0
        fill = False
        
        for item in datos:
            total += float(item['Monto'])
            
            # Formatear descripción multilínea
            desc = item['Descripción'].replace('\n', ' ')
            
            pdf.cell(anchos[0], 8, item['Fecha'], 1, 0, 'C', fill)
            pdf.cell(anchos[1], 8, item['Categoría'], 1, 0, 'C', fill)
            
            # Celda de descripción con ajuste
            x = pdf.get_x()
            y = pdf.get_y()
            pdf.multi_cell(anchos[2], 8, desc, 1, 'L', fill)
            pdf.set_xy(x + anchos[2], y)
            
            pdf.cell(anchos[3], 8, f"${float(item['Monto']):.2f}", 1, 0, 'R', fill)
            pdf.ln(8)
            fill = not fill
        
        # Fila de total
        pdf.set_font('Times', 'B', 10)
        pdf.set_fill_color(220, 230, 255)
        pdf.cell(sum(anchos[:-1]), 8, texto_total, 1, 0, 'R', True)
        pdf.cell(anchos[3], 8, f"${total:.2f}", 1, 1, 'R', True)

    def update_table(self):
        for item in self.tree.get_children():
            self.tree.delete(item)
            
        start, end = self.get_period_range()
        
        for transaction in self.transactions:
            trans_date = datetime.strptime(transaction["Fecha"], "%Y-%m-%d")
            if start <= trans_date < end:
                self.tree.insert("", tk.END, values=(
                    transaction["Fecha"],
                    transaction["Tipo"],
                    transaction["Categoría"],
                    f"${float(transaction['Monto']):.2f}",
                    transaction["Descripción"]
                ))
        
    def load_transactions(self):
        try:
            workbook = load_workbook(self.filename)
            sheet = workbook.active
            self.transactions = []
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if row:
                    self.transactions.append({
                        "Fecha": row[0],
                        "Tipo": row[1],
                        "Categoría": row[2],
                        "Monto": row[3],
                        "Descripción": row[4]
                    })
        except FileNotFoundError:
            self.transactions = []
            
    def save_transactions(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Fecha", "Tipo", "Categoría", "Monto", "Descripción"])
        
        for transaction in self.transactions:
            sheet.append([
                transaction["Fecha"],
                transaction["Tipo"],
                transaction["Categoría"],
                transaction["Monto"],
                transaction["Descripción"]
            ])
        
        workbook.save(self.filename)
        self.update_table()
    
    def get_period_range(self):
        today = datetime.today()
        if self.selected_period.get() == "Día":
            start = today.replace(hour=0, minute=0, second=0, microsecond=0)
            end = start + timedelta(days=1)
        elif self.selected_period.get() == "Semana":
            start = today - timedelta(days=today.weekday())
            start = start.replace(hour=0, minute=0, second=0, microsecond=0)
            end = start + timedelta(weeks=1)
        elif self.selected_period.get() == "Mes":
            start = today.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
            end = (start + timedelta(days=32)).replace(day=1)
        else: # Mostrar todo
            start = today.replace(year=1980, day=1, hour=0, minute=0, second=0, microsecond=0)
            end = datetime.max
        return start, end
    
    def open_add_window(self):
        self.transaction_window("Agregar Transacción")
        
    def open_edit_window(self):
        selected_item = self.tree.selection()
        if not selected_item:
            messagebox.showwarning("Advertencia", "Seleccione una transacción para editar")
            return
            
        self.selected_transaction = self.tree.item(selected_item)["values"]
        self.transaction_window("Editar Transacción", edit_mode=True)
    
    def transaction_window(self, title, edit_mode=False):
        window = Toplevel(self.root)
        window.title(title)
        window.geometry("600x500")
        window.configure(bg='#f0f0f0')
        
        main_frame = ttk.Frame(window)
        main_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=20)
        
        labels = ["Fecha (AAAA-MM-DD):", "Tipo:", "Categoría:", "Monto:", "Descripción:"]
        
        if edit_mode:
            default_values = [
                self.selected_transaction[0],
                self.selected_transaction[1],
                self.selected_transaction[2],
                self.selected_transaction[3][1:],
                self.selected_transaction[4]
            ]
        else:
            default_values = [
                datetime.today().strftime("%Y-%m-%d"),
                "Ingreso",
                "Efectivo",
                "",
                ""
            ]
        
        fields = {}
        for i, label in enumerate(labels[:-1]):
            frame = ttk.Frame(main_frame)
            frame.pack(fill=tk.X, pady=8)
            
            ttk.Label(frame, text=label).pack(side=tk.LEFT, padx=(0, 10))
            
            if label == "Tipo:":
                field = ttk.Combobox(frame, values=["Ingreso", "Gasto"], width=24)
            elif label == "Categoría:":
                field = ttk.Combobox(frame, values=["Efectivo", "Banco", "Otros"], width=24)
            else:
                field = ttk.Entry(frame, width=30)
            
            field.pack(side=tk.RIGHT, fill=tk.X, expand=True)
            field.insert(0, default_values[i])
            fields[label] = field
        
        # Campo descripción
        desc_frame = ttk.Frame(main_frame)
        desc_frame.pack(fill=tk.BOTH, expand=True, pady=8)
        
        ttk.Label(desc_frame, text=labels[-1]).pack(anchor=tk.NW, pady=(0, 5))
        desc_field = tk.Text(desc_frame, height=5, wrap=tk.WORD, bd=1, relief='flat',
                           font=('Arial', 10), bg='white', padx=8, pady=8)
        desc_field.pack(fill=tk.BOTH, expand=True)
        desc_field.insert("1.0", default_values[-1])
        fields[labels[-1]] = desc_field
        
        # Botones
        btn_frame = ttk.Frame(main_frame)
        btn_frame.pack(fill=tk.X, pady=15)
        
        ttk.Button(
            btn_frame,
            text="Guardar",
            style='Primary.TButton',
            command=lambda: self.save_transaction(fields, window, edit_mode)
        ).pack(side=tk.RIGHT, padx=5)
        
        ttk.Button(
            btn_frame,
            text="Cancelar",
            style='Secondary.TButton',
            command=window.destroy
        ).pack(side=tk.RIGHT, padx=5)
        
    def save_transaction(self, fields, window, edit_mode):
        try:
            new_transaction = {
                "Fecha": fields["Fecha (AAAA-MM-DD):"].get(),
                "Tipo": fields["Tipo:"].get(),
                "Categoría": fields["Categoría:"].get(),
                "Monto": float(fields["Monto:"].get()),
                "Descripción": fields["Descripción:"].get("1.0", tk.END).strip()
            }
            
            datetime.strptime(new_transaction["Fecha"], "%Y-%m-%d")
            
            if edit_mode:
                for i, t in enumerate(self.transactions):
                    if t["Fecha"] == self.selected_transaction[0] and t["Descripción"] == str(self.selected_transaction[4]):
                        del self.transactions[i]
                        break
            
            self.transactions.append(new_transaction)
            self.save_transactions()
            self.update_table()
            window.destroy()
            
        except ValueError as e:
            messagebox.showerror("Error", f"Dato inválido: {str(e)}")
    
    def delete_transaction(self):
        selected_item = self.tree.selection()
        if not selected_item:
            messagebox.showwarning("Advertencia", "Seleccione una transacción para eliminar")
            return
            
        if messagebox.askyesno("Confirmar", "¿Eliminar esta transacción?"):
            selected_values = self.tree.item(selected_item)["values"]

            for i, t in enumerate(self.transactions):
                print(f'i:{i}, t:{t}')
                if (t["Fecha"] == selected_values[0] and 
                    str(t["Descripción"]) == str(selected_values[4]) and
                    float(t["Monto"]) == float(selected_values[3][1:])):
                    del self.transactions[i]
                    self.save_transactions()
                    self.update_table()
                    return

            messagebox.showwarning("Advertencia", "No se pudo encontrar la transacción para eliminar")

if __name__ == "__main__":
    root = tk.Tk()
    app = FinanceApp(root)
    root.mainloop()